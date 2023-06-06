""""MASTER CALENDER"""
import sys
from openpyxl import load_workbook
import openpyxl
import openpyxl.styles


def code_list(sheet):
    """This function returns a list of course codes """
    a_1 = []
    for i in sheet['A']:
        a_1.append(i.value)
#    print(a_1)
    return a_1


def code_module(sheet):
    """This function returns a list of course modules """
    a_1 = []
    for i in sheet['B']:
        a_1.append(i.value)
    return a_1


def code_fac(sheet):
    """This function returns a list of faculty names """
    a_1 = []
    for i in sheet["C"]:
        a_1.append(i.value)
    return a_1


def code_fac2(sheet):
    """This function returns a list of faculty names """
    a_1 = []
    for i in sheet["D"]:
        a_1.append(i.value)
    return a_1


def code_fac3(sheet):
    """This function returns a list of faculty names """
    a_1 = []
    for i in sheet["E"]:
        a_1.append(i.value)
    return a_1


def update_code(code_1, sheet):
    """This function updates the excel with new course code """
    sheet.cell(row=sheet.max_row+1, column=1).value = code_1


def c_list(sheet):
    """This function returns a list of course codes in the updated excel """
    a_1 = []
    for i in sheet["A"]:
        a_1.append(i.value)
#    print(a_1)
    return a_1


def update_mod(mod, sheet):
    """This function updates the excel with new course module """
    sheet.cell(row=sheet.max_row, column=2).value = mod


def update_f(faculty_1, sheet):
    """This function updates the excel with faculty name """
    sheet.cell(row=sheet.max_row, column=3).value = faculty_1


def update_f2(faculty_2, sheet):
    """This function updates the excel with faculty name """
    sheet.cell(row=sheet.max_row, column=4).value = faculty_2


def update_f3(faculty_3, sheet):
    """This function updates the excel with faculty name """
    sheet.cell(row=sheet.max_row, column=5).value = faculty_3


def mod_list(sheet):
    """This function returns a list of course modules in the updated excel """
    a_1 = []
    for i in sheet["B"]:
        a_1.append(i.value)
    return a_1


def f_list(sheet):
    """This function returns a list of faculty in the updated excel """
    a_1 = []
    for i in sheet["C"]:
        a_1.append(i.value)
    return a_1


def list_month(sheet):
    """This function returns a list of month in the excel """
    a_1 = []
    for i in sheet["1"]:
        a_1.append(i.value)
    return a_1


def list_date(sheet, month, date, month_list):
    """This function returns a list of dates in the excel """
    a_1 = []
    col_month = month_list.index(month)+1
    for i in range(col_month, col_month + (2*date)):
        a_1.append(sheet.cell(row=2, column=i).value)
    return a_1


def update_excel(sheet, row1, col_date, program):
    """This function updates the excel depending on the conditions passed """
    if sheet.cell(row=row1, column=col_date).value is None:
        sheet.cell(row=row1, column=col_date).value = 1
        cell_to_edit = sheet.cell(row=row1, column=col_date)
        if program == 'Genesis':
            cell_to_edit.fill = openpyxl.styles.PatternFill("solid", fgColor="3498DB")
        elif program == 'StepIn':
            cell_to_edit.fill = openpyxl.styles.PatternFill("solid", fgColor="1ABC9C")

        else:
            cell_to_edit.fill = openpyxl.styles.PatternFill("solid", fgColor="9B59B6")
    else:
        sheet.cell(row=row1, column=col_date).value += 1
        cell_to_edit = sheet.cell(row=row1, column=col_date)
        cell_to_edit.fill = openpyxl.styles.PatternFill("solid", fgColor="E74C3C")


def master_loop(code_1, mod, faculty_1, faculty_2, faculty_3, l_1, l_2, l_3, l_4, l_5, sheet, month, date, slot,
                program):
    """This function get the row and column values to be updated and updates the excel sheet """
    if code_1 is not None:

        if code_1 not in l_1:
            update_code(code_1, sheet)
            l_1 = c_list(sheet)

        if mod not in l_2:
            update_mod(mod, sheet)
#        l2 = mod_list(sh)

        if faculty_1 not in l_3:
            update_f(faculty_1, sheet)
#        l3 = f_list(sh)

        if faculty_2 not in l_4:
            update_f2(faculty_2, sheet)
#        l4 = code_fac2(sh)

        if faculty_3 not in l_5:
            update_f3(faculty_3, sheet)
#        l5 = code_fac3(sh)

        row = l_1.index(code_1)+1

        month_list = list_month(sheet)

        if month not in month_list:
            sys.exit("The Value of month is not present in the calender.")

        date_list = list_date(sheet, month, date, month_list)
        col_month = month_list.index(month)+1

        if date not in date_list:
            sys.exit("The value is not present in calendar.")

        col_date = col_month + date_list.index(date)

        if slot == 'A':
            col_date += 1

        update_excel(sheet, row, col_date, program)


def master_calender(object1):
    """This is the main function """
    w_b = load_workbook('OutputFile.xlsx')
    sheet = w_b.active
    _ = sheet
    sheet = w_b['Mastercalendar']
    list_code = code_list(sheet)
    list_module = code_module(sheet)
    list_fac1 = code_fac(sheet)
    list_fac2 = code_fac2(sheet)
    list_fac3 = code_fac3(sheet)
    master_loop(object1.course_code, object1.course_name, object1.faculty1, object1.faculty2, object1.faculty3,
                list_code, list_module, list_fac1, list_fac2, list_fac3, sheet, object1.month, object1.date,
                object1.session_slot, object1.program_type)

    w_b.save("OutputFile.xlsx")


def test_code_list():
    """Testing for code list with test examples"""
    w_b = load_workbook('OutputFile.xlsx')
    sheet = w_b['Sheet2']
    test_list = code_list(sheet)
    assert test_list == ['Course Code', None, None, 101, 102, 103, 104, 105]


def test_modulename_list():
    """Testing for course list with test examples"""
    w_b = load_workbook('OutputFile.xlsx')
    sheet = w_b['Sheet2']
    test_list = code_module(sheet)
    assert test_list == [' Course Name', None, None, 'test1', 'test2', 'test3', 'test4', 'test5']
