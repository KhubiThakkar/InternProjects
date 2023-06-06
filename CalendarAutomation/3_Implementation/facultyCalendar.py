#!/usr/bin/env python3
"""The module with all functions required to generate the Faculty Calendar"""
import sys
import openpyxl
import openpyxl.styles
import pytest


def faculty_name_list(sheet):
    """To check the list of faculties already in the excel"""
    list_names = []
    for data in sheet["A"]:
        if data.value is not None:
            update_data = data.value.capitalize()
        else:
            update_data = data.value
        list_names.append(update_data)
    return list_names


def month_list(sheet):
    """Generate the list of months already in the excel"""
    list_month = []
    for data in sheet["1"]:
        list_month.append(data.value)
    return list_month


def date_list(sheet, month, date, list_month):
    """Generate the list of dates in the corresponding month"""
    listling = []
    col_month = list_month.index(month) + 1
    for cols in range(col_month, col_month + (2 * date)):
        listling.append(sheet.cell(row=2, column=cols).value)
    return listling


def update_faculty_name(name, sheet):
    """Add the faculty name if not present"""
    sheet.cell(row=sheet.max_row + 1, column=1).value = name
    return 0


def update_excel(sheet, row1, col_date, program):
    """Block the calendar"""
    if sheet.cell(row=row1, column=col_date).value is None:
        sheet.cell(row=row1, column=col_date).value = 1
        cell_to_edit = sheet.cell(row=row1, column=col_date)
        if program == 'Genesis':
            cell_to_edit.fill = openpyxl.styles.PatternFill("solid", fgColor="3498DB")
        elif program == 'StepIn':
            cell_to_edit.fill = openpyxl.styles.PatternFill("solid", fgColor="1ABC9C")
        # for any other reasons if calendar is blocked
        else:
            cell_to_edit.fill = openpyxl.styles.PatternFill("solid", fgColor="9B59B6")
    else:
        sheet.cell(row=row1, column=col_date).value += 1
        cell_to_edit = sheet.cell(row=row1, column=col_date)
        cell_to_edit.fill = openpyxl.styles.PatternFill("solid", fgColor="E74C3C")


def iterate_faculty(faculty_name, list_names, sheet, month, date, slot, program):
    """activities to perform if the faculty name is valid"""
    # if the name is not present, add the name on the last row
    if faculty_name is not None:
        if faculty_name.capitalize() not in list(map(lambda x: x.capitalize(), list(filter(lambda x: x is not None, list_names)))):
            update_faculty_name(faculty_name.capitalize(), sheet)

        # list of names in column 1
        list_names = faculty_name_list(sheet)

        # row in which the name of faculty is present
        row1 = list_names.index(faculty_name.capitalize()) + 1

        # list of all months present in the calendar
        list_month = month_list(sheet)

        # check if month is not in the calendar then quit
        if month not in list_month:
            sys.exit("The Value of month is not present in the calendar.")

        # make a list of all the dates present in the month in calendar
        list_date = date_list(sheet, month, date, list_month)
        col_month = list_month.index(month) + 1

        # check if date is in the list or not
        if date not in list_date:
            sys.exit("The value of date is not present in the calendar.")

        # get the column in which date is present for respective month
        col_date = col_month + list_date.index(date)

        # set the column according to the session slot
        if slot == 'A':
            col_date += 1

        # at this point we have row and column of cells, where we need to update the colour

        # now we are writing the numbers and adding colours to the cell
        # For faculty 1
        update_excel(sheet, row1, col_date, program)


def faculty_calendar(object1):
    """Main function used to update the Faculty Calendar"""
    filename = "OutputFile.xlsx"
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook['FacultyCalendar']

    # checking which names are present in the first column and returning a list
    list_names = faculty_name_list(sheet)

    iterate_faculty(object1.faculty1, list_names, sheet, object1.month, object1.date, object1.session_slot, object1.program_type)
    iterate_faculty(object1.faculty2, list_names, sheet, object1.month, object1.date, object1.session_slot, object1.program_type)
    iterate_faculty(object1.faculty3, list_names, sheet, object1.month, object1.date, object1.session_slot, object1.program_type)

    workbook.save(filename)


def test_date_list():
    """Testing for function: date_list"""
    filename = "OutputFile.xlsx"
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook['FacultyCalendar']
    listm = month_list(sheet)
    assert date_list(sheet, 'Jul', 2, listm) == [1, 1, 2, 2]
    listm = month_list(sheet)
    assert date_list(sheet, 'Aug', 6, listm) == [1, 1, 2, 2, 3, 3, 4, 4, 5, 5, 6, 6]


def test_update_faculty_name():
    """Testing for function: update_faculty_name"""
    filename = "OutputFile.xlsx"
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook['FacultyCalendar']
    assert update_faculty_name('sharan', sheet) == 0
