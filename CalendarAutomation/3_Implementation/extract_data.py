""" Module for extraction of data from Calender
    Run this module to execute project"""
import os
import sys
import datetime
from openpyxl import load_workbook
import management_class
import upload_file
from facultyCalendar import faculty_calendar
from mastercalend import master_calender

#Extracts Browsed file, Creates a copy in working directory
# pylint: disable=line-too-long
if upload_file.sourceFiles_list is None:
    upload_file.tk.messagebox.showerror("ProgramManagement", "Error reading file")
    sys.exit()

Calender_workBook = load_workbook(os.path.join(os.getcwd(), os.path.basename(upload_file.sourceFiles_list[0])))

#log file for debugging
with open((os.path.join(os.getcwd(), "logfile.txt")), 'a') as f:
    for CurrentSheet in Calender_workBook.sheetnames:
        Calender_workSheet = Calender_workBook[CurrentSheet]
        for RowNum, RowValue in enumerate([[data.value for data in row] for row in Calender_workSheet.rows][1:]):
            CalenderData = management_class.ProgramManagement(RowValue[0], RowValue[1], RowValue[2], RowValue[3], RowValue[4], RowValue[5], RowValue[6], RowValue[7], CurrentSheet)
            if(RowValue[0] is not None and RowValue[1] is not None and RowValue[7] is not None and (RowValue[4] is not None or RowValue[5] is not None or RowValue[6] is not None)):
                faculty_calendar(CalenderData)
                timeupdate = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                TEMP_DATA = "{} , {} , faculty calender updated for row{}\n".format(timeupdate, CurrentSheet, RowNum+2)
                f.write(TEMP_DATA)
                if RowValue[2] is not None:
                    master_calender(CalenderData)
                    timeupdate = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    TEMP_DATA = "{} , {} , Master calender updated for row{}\n".format(timeupdate, CurrentSheet, RowNum+2)
                    f.write(TEMP_DATA)
            else:
                if RowValue[0] is None:
                    timeupdate = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    TEMP_DATA = "{} , {} , Missing date value in row{}\n".format(timeupdate, CurrentSheet, RowNum+2)
                    f.write(TEMP_DATA)
                if RowValue[1] is None:
                    timeupdate = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    TEMP_DATA = "{} , {} , Missing Month value in row{}\n".format(timeupdate, CurrentSheet, RowNum+2)
                    f.write(TEMP_DATA)
                if RowValue[2] is None:
                    timeupdate = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    TEMP_DATA = "{} , {} , Missing Course code in row{}\n".format(timeupdate, CurrentSheet, RowNum+2)
                    f.write(TEMP_DATA)
                if RowValue[7] is None:
                    timeupdate = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    TEMP_DATA = "{} , {} , Missing session timing in row{}\n".format(timeupdate, CurrentSheet, RowNum+2)
                    f.write(TEMP_DATA)
                if RowValue[4] is None and RowValue[5] is None or RowValue[6] is None:
                    timeupdate = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    TEMP_DATA = "{} , {} , Missing faculty name in row{}\n".format(timeupdate, CurrentSheet, RowNum+2)
                    f.write(TEMP_DATA)
        TEMP_DATA = "                               End of {} sheet             \n".format(CurrentSheet)
        f.write(TEMP_DATA)
    f.write("---------------------------------------------------------------------------------------\n")
